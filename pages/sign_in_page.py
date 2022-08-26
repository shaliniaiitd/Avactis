from selenium.webdriver.common.by import By
from Selenium.Avactis_pytest.pages.base_page import BasePage
from Selenium.Avactis_pytest.Config.config import TestData as TD
from ddt import data, unpack

class SignInPage(BasePage):

    def __init__(self, driver):
            super().__init__(driver)
            self.driver.get(TD.url)

            self.register = (By.XPATH, "//a[@href='http://localhost/Avactis/register.php']")
            self.sign_in = (By.CSS_SELECTOR, "input[value='Sign In']")

            self.email_id = (By.NAME, "email")

            self.passwd = (By.CSS_SELECTOR, "input[name= 'passwd']")

            import openpyxl as ox

            path = "D://WebDriver/Selenium/g_page_object_model/avactis/Avactis.xlsx"

            workbook = ox.load_workbook(path)
            names = workbook.sheetnames
            sheets = workbook.worksheets
            for sheet in sheets:
                rows = sheet.max_row

                self.test_data = []
                for row in sheet.iter_rows(min_row=1, max_col=3, max_row=rows, values_only=True):
                    self.test_data.append(row)
    @data(TD.data)
    @unpack
    def click_sign_in(self,username,password):
        self.do_send_keys(self.email_id,username)
        self.do_send_keys(self.passwd,password)
        self.do_click(self.sign_in)
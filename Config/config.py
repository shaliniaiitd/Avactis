from selenium import webdriver
import openpyxl as ox


class TestData():
    url = "http://localhost/Avactis/"
    data = []

    # read data from excel file

    path = "D://WebDriver/Selenium/g_page_object_model/avactis/Avactis.xlsx"

    workbook = ox.load_workbook(path)
    names = workbook.sheetnames
    sheets = workbook.worksheets
    for sheet in sheets:
        rows = sheet.max_row

        for row in sheet.iter_rows(min_row=2, max_col=3, max_row=rows, values_only=True):
            data.append(row)
